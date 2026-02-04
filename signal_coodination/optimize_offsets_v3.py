#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
point-queue（無拡散・飽和放出）モデルで、3リンク4交差点(0-1-2-3)の交差点オフセット
(x1,x2,x3) を step=0.01 刻みで全探索し、車両1台当たり遅れ(s/veh)を最小化する。

ユーザ要件:
- 需要は固定。両方向同じ需要幅P（需要時間率）。
- 点キュー：停止による遅れは「キュー長の時間積分（veh*s）」として計算。
- 車両1台当たり遅れ = (総遅れ veh*s) / (需要幅に相当する台数)。
- 飽和交通流率 = sat_flow_veh_per_s = 0.5 veh/s。
- 「P の中を飽和交通流率で行く」＝
    各方向とも、上流端（0または3）の“需要時間” PC 秒だけ外部から sat_flow で流入。
- 無拡散：リンクは固定走行時間の純遅延（delay line）。
- 信号：各交差点 i の青開始が x_i*C（mod C）、青長が g_i*C。
  （x0は0固定。x1..x3を探索。）
- Excel入出力：experiment.xlsx の Sheet1 を入力として、結果を同xlsxに書き戻す。

入力( Sheet1 )列（既存）:
  路線番号, 系統速度, link(0,1), link(1,2), link(2,3),
  x0opt,x1opt,x2opt,x3opt,d0opt(s/veh),d1opt(s/veh),d2opt(s/veh),d3opt(s/veh),dopt(s/veh)

CONFIGシート（なければ自動作成）:
  Cycle_C_sec, Step (cycle fraction), P, sat_flow_veh_per_s, g0,g1,g2,g3, Routes_to_run, Speeds_to_run

出力:
- Sheet1: 各（路線,速度）行に最良解（rank1）のオフセットと遅れを上書き
- Summary_Top10: 全（路線,速度）Top10一覧
- R{route}_S{speed}: 各（路線,速度）Top10詳細

注意:
- 0.01刻みで x1,x2,x3 のみ探索（x0=0）。
- 離散時間：dt = Step*C。リンク遅延は delay_step = round(tau/dt) で丸める（Stepが細ければ十分）。
"""

import re
import heapq
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


CONFIG_SHEET="CONFIG"
SUMMARY_SHEET="Summary_Top10"

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[\[\]\*\?/\\:]', '_', name)
    return name[:31]

def build_mask(T:int, start:int, dur:int):
    """length-T bool mask with wrap. start in [0,T), dur in [0,T]."""
    if dur<=0:
        return [False]*T
    if dur>=T:
        return [True]*T
    m=[False]*T
    for k in range(dur):
        m[(start+k)%T]=True
    return m

def simulate_one_dir(T, dt, sat, P, green_masks, delay_steps, origin_node):
    """
    1方向（0->3 なら origin_node=0, 3->0 なら origin_node=3）をシミュレーション。
    green_masks: list of 4 bool arrays length T for nodes 0..3.
    delay_steps: (d01,d12,d23) in steps for physical direction 0->3.
    戻り: (delay_veh_s_per_node[4], vehicles_entered)
    """
    d01,d12,d23 = delay_steps

    # link arrival arrays (vehicles arriving at downstream at step t)
    maxd = max(d01,d12,d23)
    arr01=[0.0]*(T+maxd+5)
    arr12=[0.0]*(T+maxd+5)
    arr23=[0.0]*(T+maxd+5)

    q=[0.0,0.0,0.0,0.0]  # queues (veh)
    delay=[0.0,0.0,0.0,0.0]  # veh*s
    vehicles_entered=0.0

    # demand window indices: length P*T, starting at origin green start.
    # start index = first green index in mask (assume contiguous dur), so we locate it.
    mask0=green_masks[origin_node]
    try:
        start_idx=mask0.index(True)
    except ValueError:
        start_idx=0
    dur_idx=int(round(P*T))
    demand=build_mask(T, start_idx, dur_idx)

    for t in range(T):
        # external arrivals at origin
        if demand[t]:
            a = sat*dt
            q[origin_node] += a
            vehicles_entered += a

        # add link arrivals (direction depends)
        if origin_node==0:
            q[1] += arr01[t]
            q[2] += arr12[t]
            q[3] += arr23[t]
            order=[0,1,2,3]
        else:
            # reverse: physical links are 2->3 etc, but same delay steps used reversed
            q[2] += arr23[t]  # from 3 to 2 uses d23
            q[1] += arr12[t]  # from 2 to 1 uses d12
            q[0] += arr01[t]  # from 1 to 0 uses d01
            order=[3,2,1,0]

        # accumulate delay as integral of queue length (veh*s)
        for i in range(4):
            delay[i] += q[i]*dt

        # discharge in travel order (all nodes act in same step independently; point-queue, no blocking)
        for i in order:
            cap = sat*dt if green_masks[i][t] else 0.0
            dep = q[i] if q[i] < cap else cap
            q[i] -= dep

            # send to downstream delay line
            if origin_node==0:
                if i==0:
                    arr01[t+d01] += dep
                elif i==1:
                    arr12[t+d12] += dep
                elif i==2:
                    arr23[t+d23] += dep
                # i==3 exits
            else:
                # reverse direction: 3->2 uses d23, 2->1 uses d12, 1->0 uses d01
                if i==3:
                    arr23[t+d23] += dep
                elif i==2:
                    arr12[t+d12] += dep
                elif i==1:
                    arr01[t+d01] += dep
                # i==0 exits

    return delay, vehicles_entered

def eval_offsets_pointqueue(x1, x2, x3, tau01, tau12, tau23, C, step, P, sat, g):
    """
    両方向を回し、各交差点の総遅れ(veh*s)を合算して台数で割り、秒/台を返す。
    戻り: (d_per_node_sec_per_veh[4], d_total_sec_per_veh)
    """
    dt = step*C
    T = int(round(1/step))

    # delays in steps
    d01 = int(round(tau01/dt))
    d12 = int(round(tau12/dt))
    d23 = int(round(tau23/dt))

    # green masks for each node (x0=0 fixed)
    xs=[0.0, x1, x2, x3]
    green_masks=[]
    for i in range(4):
        start = int(round(xs[i]*T)) % T
        dur   = int(round(g[i]*T))
        green_masks.append(build_mask(T, start, dur))

    # forward 0->3
    delay_f, veh_f = simulate_one_dir(T, dt, sat, P, green_masks, (d01,d12,d23), origin_node=0)
    # backward 3->0
    delay_b, veh_b = simulate_one_dir(T, dt, sat, P, green_masks, (d01,d12,d23), origin_node=3)

    delay_total=[delay_f[i]+delay_b[i] for i in range(4)]
    veh_total = veh_f + veh_b
    if veh_total<=0:
        # avoid division by zero
        d_nodes=[float("inf")]*4
        return d_nodes, float("inf")

    d_nodes=[delay_total[i]/veh_total for i in range(4)]  # sec/veh
    return d_nodes, sum(d_nodes)

def topk_offsets(tau01, tau12, tau23, C, step, P, sat, g, k=10):
    """0<=x<1 を step 刻みで全探索し、遅れ小さい順の上位kを返す。"""
    n = int(round(1/step))
    grid = [round(i*step, 10) for i in range(n)]

    heap=[]  # (-d_total, x1,x2,x3,d0,d1,d2,d3)
    for x1 in grid:
        for x2 in grid:
            for x3 in grid:
                d_nodes, d_total = eval_offsets_pointqueue(x1,x2,x3,tau01,tau12,tau23,C,step,P,sat,g)
                item=(-d_total, x1,x2,x3, d_nodes[0],d_nodes[1],d_nodes[2],d_nodes[3])
                if len(heap)<k:
                    heapq.heappush(heap,item)
                else:
                    if item[0] > heap[0][0]:
                        heapq.heapreplace(heap,item)
    res=[(-h[0],h[1],h[2],h[3],h[4],h[5],h[6],h[7]) for h in heap]
    res.sort(key=lambda x:x[0])
    return res

def ensure_config_sheet(wb):
    if CONFIG_SHEET in wb.sheetnames:
        return wb[CONFIG_SHEET]
    ws=wb.create_sheet(CONFIG_SHEET,0)
    ws["A1"]="Cycle_C_sec"; ws["B1"]=120
    ws["A2"]="Step (cycle fraction)"; ws["B2"]=0.01
    ws["A3"]="P (demand time rate, same both dirs)"; ws["B3"]=0.5
    ws["A4"]="sat_flow_veh_per_s"; ws["B4"]=0.5
    ws["A5"]="g0 (if blank, use P)"; ws["B5"]=""  # optional
    ws["A6"]="g1"; ws["B6"]=0.6
    ws["A7"]="g2"; ws["B7"]=0.6
    ws["A8"]="g3 (if blank, use P)"; ws["B8"]=""  # optional
    ws["A10"]="Routes_to_run (comma)"; ws["B10"]="311"
    ws["A11"]="Speeds_to_run (comma or 'all')"; ws["B11"]="all"
    bold=Font(bold=True)
    for r in range(1,12):
        ws[f"A{r}"].font=bold
    ws.column_dimensions["A"].width=42
    ws.column_dimensions["B"].width=42
    ws.freeze_panes="A2"
    return ws

def read_config(cfg):
    C=float(cfg["B1"].value)
    step=float(cfg["B2"].value)
    P=float(cfg["B3"].value)
    sat=float(cfg["B4"].value)

    def cell_float(addr, default=None):
        v=cfg[addr].value
        if v is None: return default
        s=str(v).strip()
        if s=="": return default
        return float(s)

    g0 = cell_float("B5", None)
    g1 = float(cfg["B6"].value)
    g2 = float(cfg["B7"].value)
    g3 = cell_float("B8", None)
    if g0 is None: g0 = P
    if g3 is None: g3 = P
    g=[g0,g1,g2,g3]

    routes_str=str(cfg["B10"].value).strip()
    routes=[int(x) for x in re.split(r"[\s,]+", routes_str) if x]

    speeds_raw=str(cfg["B11"].value).strip().lower()
    if speeds_raw=="all":
        speeds=None
    else:
        speeds=[float(x) for x in re.split(r"[\s,]+", speeds_raw) if x]
    return C, step, P, sat, g, routes, speeds

def header_map(ws):
    headers=[ws.cell(1,c).value for c in range(1, ws.max_column+1)]
    return {h:i+1 for i,h in enumerate(headers) if h is not None}

def iter_rows(ws, col_route, col_speed):
    cur=None
    for r in range(2, ws.max_row+1):
        rn=ws.cell(r,col_route).value
        if rn is not None:
            cur=rn
        sp=ws.cell(r,col_speed).value
        if sp is None:
            continue
        yield cur, float(sp), r

def get_tau_from_row(ws, r, col_speed, col_L01, col_L12, col_L23):
    v=float(ws.cell(r,col_speed).value)  # km/h
    mps=v*1000.0/3600.0
    L01=float(ws.cell(r,col_L01).value)  # m
    L12=float(ws.cell(r,col_L12).value)
    L23=float(ws.cell(r,col_L23).value)
    return (L01/mps, L12/mps, L23/mps)

def write_best_to_main(ws, r, col, best):
    dtotal,x1,x2,x3,d0,d1,d2,d3 = best
    ws.cell(r, col["x0opt"]).value = 0.0
    ws.cell(r, col["x1opt"]).value = x1
    ws.cell(r, col["x2opt"]).value = x2
    ws.cell(r, col["x3opt"]).value = x3
    ws.cell(r, col["d0opt(s/veh)"]).value = d0
    ws.cell(r, col["d1opt(s/veh)"]).value = d1
    ws.cell(r, col["d2opt(s/veh)"]).value = d2
    ws.cell(r, col["d3opt(s/veh)"]).value = d3
    ws.cell(r, col["dopt(s/veh)"]).value = dtotal

def make_detail_sheet(wb, route, speed, top10, C, step, P, sat, g):
    title=sanitize_sheet_name(f"R{route}_S{speed:g}")
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws=wb.create_sheet(title)

    ws["A1"]="route"; ws["B1"]=route
    ws["A2"]="speed(km/h)"; ws["B2"]=speed
    ws["A3"]="C(sec)"; ws["B3"]=C
    ws["A4"]="step"; ws["B4"]=step
    ws["A5"]="P"; ws["B5"]=P
    ws["A6"]="sat_flow"; ws["B6"]=sat
    ws["A7"]="g0,g1,g2,g3"; ws["B7"]=str(g)

    ws["A9"]="rank"
    headers=["d_total(s/veh)","x1","x2","x3","d0","d1","d2","d3"]
    for j,h in enumerate(headers, start=2):
        ws.cell(9,j).value=h
    bold=Font(bold=True)
    for c in range(1,10):
        ws.cell(9,c).font=bold

    for i,row in enumerate(top10, start=1):
        dtotal,x1,x2,x3,d0,d1,d2,d3=row
        ws.cell(9+i,1).value=i
        for j,v in enumerate([dtotal,x1,x2,x3,d0,d1,d2,d3], start=2):
            ws.cell(9+i,j).value=float(v)

    ws.freeze_panes="A10"
    ws.column_dimensions["A"].width=6
    for c in "BCDEFGHI":
        ws.column_dimensions[c].width=16
    return title

def make_summary_sheet(wb, rows):
    if SUMMARY_SHEET in wb.sheetnames:
        wb.remove(wb[SUMMARY_SHEET])
    ws=wb.create_sheet(SUMMARY_SHEET,0)
    headers=["route","speed(km/h)","rank","d_total(s/veh)","x1","x2","x3","d0","d1","d2","d3","detail_sheet"]
    for j,h in enumerate(headers, start=1):
        ws.cell(1,j).value=h
        ws.cell(1,j).font=Font(bold=True)
    for i,rr in enumerate(rows, start=2):
        for j,v in enumerate(rr, start=1):
            ws.cell(i,j).value=v
    ws.freeze_panes="A2"
    for j in range(1,len(headers)+1):
        ws.column_dimensions[get_column_letter(j)].width=16
    ws.column_dimensions["L"].width=22
    return ws

def main(xlsx_path: str, out_path: str | None = None):
    wb=openpyxl.load_workbook(xlsx_path)
    cfg=ensure_config_sheet(wb)
    C, step, P, sat, g, target_routes, target_speeds = read_config(cfg)

    ws=wb["Sheet1"]
    col=header_map(ws)
    required=["路線番号","系統速度","link(0,1)","link(1,2)","link(2,3)",
              "x0opt","x1opt","x2opt","x3opt","d0opt(s/veh)","d1opt(s/veh)","d2opt(s/veh)","d3opt(s/veh)","dopt(s/veh)"]
    missing=[h for h in required if h not in col]
    if missing:
        raise ValueError(f"Sheet1 に必要な見出しがありません: {missing}")

    summary_rows=[]

    for route, speed, r in iter_rows(ws, col["路線番号"], col["系統速度"]):
        if route not in target_routes:
            continue
        if target_speeds is not None and speed not in target_speeds:
            continue

        tau01,tau12,tau23=get_tau_from_row(ws,r,col["系統速度"],col["link(0,1)"],col["link(1,2)"],col["link(2,3)"])
        top10=topk_offsets(tau01,tau12,tau23,C,step,P,sat,g,k=10)

        # best to Sheet1
        write_best_to_main(ws,r,col,top10[0])

        detail=make_detail_sheet(wb,route,speed,top10,C,step,P,sat,g)

        for rank,row in enumerate(top10, start=1):
            dtotal,x1,x2,x3,d0,d1,d2,d3=row
            summary_rows.append([route,speed,rank,dtotal,x1,x2,x3,d0,d1,d2,d3,detail])

    make_summary_sheet(wb, summary_rows)

    if out_path is None:
        p=Path(xlsx_path)
        out_path=str(p.with_name(p.stem+"_out.xlsx"))

    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__=="__main__":
    import argparse
    ap=argparse.ArgumentParser()
    ap.add_argument("xlsx", help="input xlsx (e.g. experiment.xlsx)")
    ap.add_argument("--out", default=None, help="output xlsx (default: input_stem_out.xlsx)")
    args=ap.parse_args()
    main(args.xlsx, args.out)
