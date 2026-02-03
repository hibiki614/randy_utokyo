import pandas as pd
import matplotlib.pyplot as plt
import math
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# ===== file paths =====
road = "failure_data.xlsx"


# ===== read road characteristics =====

wb = load_workbook(filename = road, read_only = True, data_only = True)
ws = wb.active

cycle = int(ws['A2'].value)

intersections = dict([(k, dict(green_split = float(ws['D' + str(k + 2)].value))) for k in range(15)])
links = dict([(k, dict(length = int(ws['G' + str(k + 2)].value), left = int(ws['H' + str(k + 2)].value), right = int(ws['I' + str(k + 2)].value), speed = float(ws['R' + str(k + 2)].value), x = float(ws['M' + str(k + 2)].value))) for k in range(14)])


intersections[0]['position'] = 0
intersections[0]['x'] = 0
for k in range(14) :
    link = links[k]
    intersections[k+1]['position'] = intersections[k]['position'] + link['length']
    intersections[k+1]['x'] = (0*intersections[k]['x'] + link['x']) % 1
    link['dttemp'] = link['length'] / link['speed']




print('\n'.join([str(links[k]['speed']) for k in range(14)]))
#exit()
left_stream_dts = [cycle * intersections[0]['x']] + [links[k]['dttemp'] for k in range(14)]
left_stream_first = [sum(left_stream_dts[:k]) % cycle for k in range(1,16)]
left_stream_last = [(cycle * intersections[0]['green_split'] + sum(left_stream_dts[:k])) % cycle for k in range(1,16)]
right_stream_dts = [cycle * (1 + intersections[0]['x'])] + [- links[k]['dttemp'] for k in range(14)]
right_stream_first = [sum(right_stream_dts[:k]) % cycle for k in range(1,16)]
right_stream_last = [(cycle * intersections[0]['green_split'] + sum(right_stream_dts[:k])) % cycle for k in range(1,16)]
nC=3
fig, ax = plt.subplots(figsize=(16, 9))
for k in range(15) :
    intersection = intersections[k]
    ax.plot([intersection['position']] * 2, [[cycle * (intersection['x'] + i) for i in range(-1, nC)] , [cycle * (intersection['x'] + intersection['green_split'] + i) for i in range(-1, nC)]], 'g:')
    ax.plot([intersection['position']] * 2, [[cycle * (intersection['x'] + intersection['green_split'] + i) for i in range(-1, nC)] , [cycle * (intersection['x'] + i + 1) for i in range(-1, nC)]], 'r-')
    if k < 14 :
        next_intersection = intersections[k+1]
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + left_stream_first[k] for i in range(nC)] , [cycle * i + left_stream_first[k] + links[k]['dttemp'] for i in range(nC)]], 'k')
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + left_stream_last[k] for i in range(nC)] , [cycle * i + left_stream_last[k] + links[k]['dttemp'] for i in range(nC)]], 'k', alpha=.5)
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + right_stream_first[k] for i in range(nC)] , [cycle * i + right_stream_first[k] - links[k]['dttemp'] for i in range(nC)]], 'b')
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + right_stream_last[k] for i in range(nC)] , [cycle * i + right_stream_last[k] - links[k]['dttemp'] for i in range(nC)]], 'b', alpha=.5)
plt.tight_layout()
plt.show()
# exit()

