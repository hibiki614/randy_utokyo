#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""optimize_offsets_v7.py  (continuous-time fluid queue / cumulative-curve-equivalent)

- 3リンク(0-1-2-3)のオフセット x1,x2,x3 を 0.01(サイクル比)刻みで総当たり
- point-queue + 飽和流率 sat の流体近似 (FIFO) を連続時間(イベント駆動)で解く
- 車群の「分裂・再分裂」は、信号による周期的サービスで自動的に表現される
- 遅れは「停止量の総和」= ∑∫Q(t)dt (veh*s) を 総台数で割り s/veh を出力

実行例:
  python optimize_offsets_v7.py experiment.xlsx --out experiment_out.xlsx
"""

import re
import heapq
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

CONFIG_SHEET = "CONFIG"
SUMMARY_SHEET = "Summary_Top10"

REQUIRED_HEADERS = [
    "路線番号", "系統速度",
    "link(0,1)", "link(1,2)", "link(2,3)",
    "x0opt", "x1opt", "x2opt", "x3opt",
    "d0opt(s/veh)", "d1opt(s/veh)", "d2opt(s/veh)", "d3opt(s/veh)", "dopt(s/veh)"
]

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[\[\]\*\?/\\:]', '_', name)
    return name[:31]

def header_map(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if h is not None}

def find_data_sheet(wb, preferred_name=None):
    if preferred_name:
        if preferred_name not in wb.sheetnames:
            raise ValueError(f'--sheet で指定した "{preferred_name}" が見つかりません。')
        ws = wb[preferred_name]
        col = header_map(ws)
        missing = [h for h in REQUIRED_HEADERS if h not in col]
        if missing:
            raise ValueError(f'"{preferred_name}" に必要な見出しがありません: {missing}')
        return ws

    for name in wb.sheetnames:
        ws = wb[name]
        col = header_map(ws)
        missing = [h for h in REQUIRED_HEADERS if h not in col]
        if not missing:
            return ws
    raise ValueError(f"必要な見出し {REQUIRED_HEADERS} をすべて含むシートが見つかりません。")

def ensure_config_sheet(wb):
    if CONFIG_SHEET in wb.sheetnames:
        return wb[CONFIG_SHEET]
    ws = wb.create_sheet(CONFIG_SHEET, 0)
    ws["A1"]  = "Cycle_C_sec"; ws["B1"]  = 120
    ws["A2"]  = "Step (cycle fraction)"; ws["B2"]  = 0.01
    ws["A3"]  = "P (demand time rate, same both dirs)"; ws["B3"]  = 0.4
    ws["A4"]  = "sat_flow_veh_per_s"; ws["B4"]  = 0.5
    ws["A5"]  = "g0 (if blank, use P)"; ws["B5"]  = ""
    ws["A6"]  = "g1"; ws["B6"]  = 0.6
    ws["A7"]  = "g2"; ws["B7"]  = 0.6
    ws["A8"]  = "g3 (if blank, use P)"; ws["B8"]  = ""
    ws["A9"]  = "Routes_to_run (comma or 'all')"; ws["B9"]  = "all"
    ws["A10"] = "Speeds_to_run (comma or 'all')"; ws["B10"] = "all"
    bold = Font(bold=True)
    for r in range(1, 13):
        ws[f"A{r}"].font = bold
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "A2"
    return ws

def _parse_list_or_all(value, kind="int"):
    if value is None:
        return None
    s = str(value).strip().lower()
    if s == "" or s == "all":
        return None
    parts = [p for p in re.split(r"[\s,]+", s) if p]
    if kind == "int":
        try:
            return [int(p) for p in parts]
        except ValueError as e:
            raise ValueError(f"Routes_to_run は整数のカンマ区切りか 'all' で指定してください。いま: {value!r}") from e
    else:
        try:
            return [float(p) for p in parts]
        except ValueError as e:
            raise ValueError(f"Speeds_to_run は数値のカンマ区切りか 'all' で指定してください。いま: {value!r}") from e

def read_config(cfg):
    kv = {}
    for r in range(1, cfg.max_row + 1):
        k = cfg.cell(r, 1).value
        v = cfg.cell(r, 2).value
        if k is None:
            continue
        ks = str(k).strip()
        if ks:
            kv[ks] = v

    def must_float(key):
        v = kv.get(key, None)
        if v is None or str(v).strip() == "":
            raise ValueError(f"CONFIG の '{key}' が空です。")
        return float(v)

    def opt_float(key):
        v = kv.get(key, None)
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        return float(s)

    C    = must_float("Cycle_C_sec")
    step = must_float("Step (cycle fraction)")
    P    = must_float("P (demand time rate, same both dirs)")
    sat  = must_float("sat_flow_veh_per_s")

    g0 = opt_float("g0 (if blank, use P)")
    g1 = must_float("g1")
    g2 = must_float("g2")
    g3 = opt_float("g3 (if blank, use P)")
    if g0 is None: g0 = P
    if g3 is None: g3 = P
    g = [g0, g1, g2, g3]

    routes = _parse_list_or_all(kv.get("Routes_to_run (comma or 'all')", kv.get("Routes_to_run (comma)", None)), kind="int")
    speeds = _parse_list_or_all(kv.get("Speeds_to_run (comma or 'all')", None), kind="float")

    # validation
    if not (C > 0):
        raise ValueError("Cycle_C_sec は正の値にしてください。")
    if not (0 < step <= 1):
        raise ValueError("Step は (0,1] の範囲にしてください。例: 0.01")
    inv = 1 / step
    if abs(inv - round(inv)) > 1e-9:
        raise ValueError("Step は 1/Step が整数になる値にしてください。例: 0.01, 0.02, 0.005")
    if not (0 < P <= 1):
        raise ValueError("P は (0,1] の範囲にしてください。")
    if not (sat > 0):
        raise ValueError("sat_flow_veh_per_s は正の値にしてください。")
    for i, gi in enumerate(g):
        if not (0 < gi <= 1):
            raise ValueError(f"g{i} は (0,1] の範囲にしてください。いま g{i}={gi}")
    return C, step, P, sat, g, routes, speeds

def iter_rows(ws, col_route, col_speed):
    cur = None
    for r in range(2, ws.max_row + 1):
        rn = ws.cell(r, col_route).value
        if rn is not None:
            cur = rn
        sp = ws.cell(r, col_speed).value
        if sp is None:
            continue
        yield cur, float(sp), r

def get_tau_from_row(ws, r, col_speed, col_L01, col_L12, col_L23):
    v = float(ws.cell(r, col_speed).value)  # km/h
    mps = v * 1000.0 / 3600.0
    L01 = float(ws.cell(r, col_L01).value)  # m
    L12 = float(ws.cell(r, col_L12).value)
    L23 = float(ws.cell(r, col_L23).value)
    return (L01 / mps, L12 / mps, L23 / mps)

# ---- signal / queue primitives ----

def signal_state_and_next_switch(t: float, C: float, x: float, g: float):
    gs = x * C
    gd = g * C
    phase = (t - gs) % C
    if phase < gd:
        return True, t + (gd - phase)  # green -> ends
    else:
        return False, t + (C - phase)  # red -> next green start

def service_rate(t: float, C: float, x: float, g: float, sat: float) -> float:
    green, _ = signal_state_and_next_switch(t, C, x, g)
    return sat if green else 0.0

def depart_rate(q: float, a: float, s: float) -> float:
    # FIFO fluid point queue
    if q > 1e-12:
        return s
    return min(a, s)

class ArrivalSchedule:
    def __init__(self):
        self.h = []  # (time, new_rate)
    def push(self, t, rate):
        heapq.heappush(self.h, (t, rate))
    def peek_time(self):
        return self.h[0][0] if self.h else None
    def pop_all_at(self, t, eps=1e-9):
        changed = False
        last_rate = None
        while self.h and abs(self.h[0][0] - t) <= eps:
            _, last_rate = heapq.heappop(self.h)
            changed = True
        return changed, last_rate

def simulate_direction(path_nodes, taus, C, P, sat, xs, g, t_start_demand):
    demand_T = P * C
    t_end_demand = t_start_demand + demand_T
    N_dir = sat * demand_T

    q = {n: 0.0 for n in path_nodes}
    a = {n: 0.0 for n in path_nodes}
    s = {n: 0.0 for n in path_nodes}
    d = {n: 0.0 for n in path_nodes}
    delay = {n: 0.0 for n in path_nodes}

    sched = {n: ArrivalSchedule() for n in path_nodes}

    src = path_nodes[0]
    t = 0.0

    next_switch = {}
    for n in path_nodes:
        s[n] = service_rate(t, C, xs[n], g[n], sat)
        _, ns = signal_state_and_next_switch(t, C, xs[n], g[n])
        next_switch[n] = ns

    a[src] = sat if (t_start_demand <= t < t_end_demand) else 0.0
    for n in path_nodes:
        d[n] = depart_rate(q[n], a[n], s[n])

    for i, n in enumerate(path_nodes[:-1]):
        if d[n] != 0.0:
            sched[path_nodes[i+1]].push(t + taus[i], d[n])

    out_cum = 0.0
    last_signal = path_nodes[-1]
    finish_time = None

    def next_empty_time(n, t_now):
        if q[n] <= 1e-12:
            return None
        net = a[n] - d[n]
        if net < -1e-12:
            return t_now + q[n] / (-net)
        return None

    max_events = 500000
    ev = 0
    while ev < max_events:
        ev += 1

        if out_cum >= N_dir - 1e-9:
            finish_time = t + taus[-1]
            break

        times = []
        for n in path_nodes:
            times.append(next_switch[n])
        if t < t_start_demand - 1e-9:
            times.append(t_start_demand)
        if t < t_end_demand - 1e-9:
            times.append(t_end_demand)
        for n in path_nodes[1:]:
            pt = sched[n].peek_time()
            if pt is not None:
                times.append(pt)
        for n in path_nodes:
            et = next_empty_time(n, t)
            if et is not None:
                times.append(et)

        t_next = min(times)
        if t_next < t + 1e-12:
            t_next = t
        dt = t_next - t

        if dt > 0:
            # integrate
            for n in path_nodes:
                net = a[n] - d[n]
                delay[n] += q[n] * dt + 0.5 * net * dt * dt
                q[n] += net * dt
                if q[n] < 0:
                    q[n] = 0.0

            out_cum += d[last_signal] * dt
            if out_cum >= N_dir - 1e-9:
                excess = out_cum - N_dir
                if d[last_signal] > 1e-12:
                    dt_back = excess / d[last_signal]
                    t_cross = t_next - dt_back
                else:
                    t_cross = t_next
                finish_time = t_cross + taus[-1]
                out_cum = N_dir
                t = t_next
                break

        t = t_next

        # events at t
        for n in path_nodes:
            if abs(next_switch[n] - t) <= 1e-9:
                s[n] = service_rate(t, C, xs[n], g[n], sat)
                _, ns = signal_state_and_next_switch(t, C, xs[n], g[n])
                next_switch[n] = ns

        if abs(t - t_start_demand) <= 1e-9:
            a[src] = sat
        if abs(t - t_end_demand) <= 1e-9:
            a[src] = 0.0

        for n in path_nodes[1:]:
            changed, rate = sched[n].pop_all_at(t)
            if changed:
                a[n] = rate

        for idx, n in enumerate(path_nodes):
            old = d[n]
            d[n] = depart_rate(q[n], a[n], s[n])
            if abs(d[n] - old) > 1e-12 and idx < len(path_nodes) - 1:
                sched[path_nodes[idx+1]].push(t + taus[idx], d[n])

    if finish_time is None:
        finish_time = t + taus[-1]
    return delay, N_dir, finish_time

def evaluate_offsets(tau01, tau12, tau23, C, step, P, sat, g, x1, x2, x3):
    xs = {0: 0.0, 1: x1, 2: x2, 3: x3}

    delay_f, N_f, T_f = simulate_direction([0,1,2], [tau01, tau12, tau23], C, P, sat, xs, g, t_start_demand=xs[0]*C)
    delay_b, N_b, T_b = simulate_direction([3,2,1], [tau23, tau12, tau01], C, P, sat, xs, g, t_start_demand=xs[3]*C)

    N_total = N_f + N_b  # = 2*sat*P*C
    D = {0:0.0,1:0.0,2:0.0,3:0.0}
    for n,v in delay_f.items():
        D[n] += v
    for n,v in delay_b.items():
        D[n] += v

    d_node = [D[i] / N_total for i in range(4)]
    d_total = sum(d_node)
    return d_node, d_total, (T_f, T_b), N_total

def topk_offsets(tau01, tau12, tau23, C, step, P, sat, g, k=10, progress_every=200000):
    n = int(round(1 / step))
    grid = [round(i * step, 10) for i in range(n)]
    total_iters = n * n * n
    heap = []
    cnt = 0

    for x1 in grid:
        for x2 in grid:
            for x3 in grid:
                cnt += 1
                if progress_every and (cnt == 1 or cnt % progress_every == 0 or cnt == total_iters):
                    pct = 100.0 * cnt / total_iters
                    print(f"    search: {cnt:,}/{total_iters:,} ({pct:.1f}%)", flush=True)

                d_node, d_total, _, _ = evaluate_offsets(tau01, tau12, tau23, C, step, P, sat, g, x1, x2, x3)
                item = (-d_total, x1, x2, x3, d_node[0], d_node[1], d_node[2], d_node[3])
                if len(heap) < k:
                    heapq.heappush(heap, item)
                else:
                    if item[0] > heap[0][0]:
                        heapq.heapreplace(heap, item)

    res = [(-h[0], h[1], h[2], h[3], h[4], h[5], h[6], h[7]) for h in heap]
    res.sort(key=lambda x: x[0])
    return res

def write_best_to_main(ws, r, col, best):
    dtotal, x1, x2, x3, d0, d1, d2, d3 = best
    ws.cell(r, col["x0opt"]).value = 0.0
    ws.cell(r, col["x1opt"]).value = x1
    ws.cell(r, col["x2opt"]).value = x2
    ws.cell(r, col["x3opt"]).value = x3
    ws.cell(r, col["d0opt(s/veh)"]).value = d0
    ws.cell(r, col["d1opt(s/veh)"]).value = d1
    ws.cell(r, col["d2opt(s/veh)"]).value = d2
    ws.cell(r, col["d3opt(s/veh)"]).value = d3
    ws.cell(r, col["dopt(s/veh)"]).value = dtotal

def make_detail_sheet(wb, route, speed, top10, C, step, P, sat, g):
    title = sanitize_sheet_name(f"R{route}_S{speed:g}")
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws = wb.create_sheet(title)

    ws["A1"] = "route"; ws["B1"] = route
    ws["A2"] = "speed(km/h)"; ws["B2"] = speed
    ws["A3"] = "C(sec)"; ws["B3"] = C
    ws["A4"] = "step(x-grid)"; ws["B4"] = step
    ws["A5"] = "P"; ws["B5"] = P
    ws["A6"] = "sat_flow"; ws["B6"] = sat
    ws["A7"] = "g0,g1,g2,g3"; ws["B7"] = str(g)

    ws["A9"] = "rank"
    headers = ["d_total(s/veh)", "x1", "x2", "x3", "d0", "d1", "d2", "d3"]
    for j, h in enumerate(headers, start=2):
        ws.cell(9, j).value = h
    bold = Font(bold=True)
    for c in range(1, 10):
        ws.cell(9, c).font = bold

    for i, row in enumerate(top10, start=1):
        dtotal, x1, x2, x3, d0, d1, d2, d3 = row
        ws.cell(9 + i, 1).value = i
        for j, v in enumerate([dtotal, x1, x2, x3, d0, d1, d2, d3], start=2):
            ws.cell(9 + i, j).value = float(v)

    ws.freeze_panes = "A10"
    ws.column_dimensions["A"].width = 6
    for c in "BCDEFGHI":
        ws.column_dimensions[c].width = 16
    return title

def make_summary_sheet(wb, rows):
    if SUMMARY_SHEET in wb.sheetnames:
        wb.remove(wb[SUMMARY_SHEET])
    ws = wb.create_sheet(SUMMARY_SHEET, 0)
    headers = ["route", "speed(km/h)", "rank", "d_total(s/veh)", "x1", "x2", "x3", "d0", "d1", "d2", "d3", "detail_sheet"]
    for j, h in enumerate(headers, start=1):
        ws.cell(1, j).value = h
        ws.cell(1, j).font = Font(bold=True)
    for i, rr in enumerate(rows, start=2):
        for j, v in enumerate(rr, start=1):
            ws.cell(i, j).value = v
    ws.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.column_dimensions["L"].width = 22
    return ws

def main(xlsx_path: str, out_path: str | None = None, sheet_name: str | None = None, progress_every: int = 200000):
    wb = openpyxl.load_workbook(xlsx_path)
    cfg = ensure_config_sheet(wb)
    C, step, P, sat, g, target_routes, target_speeds = read_config(cfg)

    ws = find_data_sheet(wb, preferred_name=sheet_name)
    col = header_map(ws)

    candidates = []
    for route, speed, r in iter_rows(ws, col["路線番号"], col["系統速度"]):
        if route is None:
            continue
        if target_routes is not None and int(route) not in target_routes:
            continue
        if target_speeds is not None and float(speed) not in target_speeds:
            continue
        candidates.append((int(route), float(speed), r))

    total_cases = len(candidates)
    print(f"CONFIG loaded: C={C}, step={step}, P={P}, sat={sat}, g={g}", flush=True)
    print(f"Using sheet: {ws.title}", flush=True)
    print(f"Total cases to run: {total_cases}", flush=True)

    summary_rows = []

    for idx, (route, speed, r) in enumerate(candidates, start=1):
        print(f"\nCase {idx}/{total_cases}: route={route} speed={speed:g}", flush=True)
        tau01, tau12, tau23 = get_tau_from_row(ws, r, col["系統速度"], col["link(0,1)"], col["link(1,2)"], col["link(2,3)"])

        top10 = topk_offsets(tau01, tau12, tau23, C, step, P, sat, g, k=10, progress_every=progress_every)

        write_best_to_main(ws, r, col, top10[0])
        detail = make_detail_sheet(wb, route, speed, top10, C, step, P, sat, g)

        for rank, row in enumerate(top10, start=1):
            dtotal, x1, x2, x3, d0, d1, d2, d3 = row
            summary_rows.append([route, speed, rank, dtotal, x1, x2, x3, d0, d1, d2, d3, detail])

        print(f"done: route={route} speed={speed:g} best_d={top10[0][0]:.6g}", flush=True)

    make_summary_sheet(wb, summary_rows)

    if out_path is None:
        p = Path(xlsx_path)
        out_path = str(p.with_name(p.stem + "_out.xlsx"))

    wb.save(out_path)
    print(f"\nSaved: {out_path}", flush=True)

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="input xlsx (e.g. experiment.xlsx)")
    ap.add_argument("--out", default=None, help="output xlsx (default: input_stem_out.xlsx)")
    ap.add_argument("--sheet", default=None, help="data sheet name (optional). If omitted, auto-detect.")
    ap.add_argument("--progress-every", type=int, default=200000,
                    help="print inner-loop progress every N combos (default: 200000). Set 0 to disable.")
    args = ap.parse_args()
    main(args.xlsx, args.out, args.sheet, progress_every=args.progress_every)
