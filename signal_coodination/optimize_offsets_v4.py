#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
optimize_offsets_v4.py

改善点（v3→v4）
- CONFIGの Routes_to_run / Speeds_to_run に "all" を指定できるように修正（今回のエラー原因）
- CONFIG値のバリデーションと、分かりやすいエラーメッセージ
- 入力シート名が "Sheet1" でなくても、必要な見出しが揃っている最初のシートを自動検出
  （必要なら --sheet で指定も可）

モデル・仕様（v3同様）
- point-queue（無拡散・純遅延リンク）
- 両方向同需要：需要時間率 P、需要窓 PC 秒だけ外部から sat_flow で流入（矩形波）
- 飽和交通流率 sat_flow_veh_per_s = 0.5 veh/s（CONFIGで変更可）
- 信号：交差点 i の青開始が x_i*C（mod C）、青長が g_i*C
- x0=0固定、探索変数は x1,x2,x3 を step=0.01刻みで全探索
- 遅れ：キュー長の時間積分（veh*s）を台数で割って s/veh
- Excel：入力xlsxに対し、最良解を元シートに書き込み + Top10の一覧/詳細シート追加
"""

import re
import heapq
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

CONFIG_SHEET = "CONFIG"
SUMMARY_SHEET = "Summary_Top10"

REQUIRED_HEADERS = [
    "路線番号", "系統速度",
    "link(0,1)", "link(1,2)", "link(2,3)",
    "x0opt", "x1opt", "x2opt", "x3opt",
    "d0opt(s/veh)", "d1opt(s/veh)", "d2opt(s/veh)", "d3opt(s/veh)", "dopt(s/veh)"
]

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[\[\]\*\?/\\:]', '_', name)
    return name[:31]

def build_mask(T: int, start: int, dur: int):
    """length-T bool mask with wrap. start in [0,T), dur in [0,T]."""
    if dur <= 0:
        return [False] * T
    if dur >= T:
        return [True] * T
    m = [False] * T
    for k in range(dur):
        m[(start + k) % T] = True
    return m

def header_map(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if h is not None}

def find_data_sheet(wb, preferred_name=None):
    """必要ヘッダが揃うシートを返す。"""
    if preferred_name:
        if preferred_name not in wb.sheetnames:
            raise ValueError(f'--sheet で指定した "{preferred_name}" が見つかりません。')
        ws = wb[preferred_name]
        col = header_map(ws)
        missing = [h for h in REQUIRED_HEADERS if h not in col]
        if missing:
            raise ValueError(f'"{preferred_name}" に必要な見出しがありません: {missing}')
        return ws

    # 自動検出
    for name in wb.sheetnames:
        ws = wb[name]
        col = header_map(ws)
        missing = [h for h in REQUIRED_HEADERS if h not in col]
        if not missing:
            return ws
    raise ValueError(f"必要な見出し {REQUIRED_HEADERS} をすべて含むシートが見つかりません。")

def ensure_config_sheet(wb):
    if CONFIG_SHEET in wb.sheetnames:
        return wb[CONFIG_SHEET]
    ws = wb.create_sheet(CONFIG_SHEET, 0)
    ws["A1"] = "Cycle_C_sec"; ws["B1"] = 120
    ws["A2"] = "Step (cycle fraction)"; ws["B2"] = 0.01
    ws["A3"] = "P (demand time rate, same both dirs)"; ws["B3"] = 0.5
    ws["A4"] = "sat_flow_veh_per_s"; ws["B4"] = 0.5
    ws["A5"] = "g0 (if blank, use P)"; ws["B5"] = ""
    ws["A6"] = "g1"; ws["B6"] = 0.6
    ws["A7"] = "g2"; ws["B7"] = 0.6
    ws["A8"] = "g3 (if blank, use P)"; ws["B8"] = ""
    ws["A10"] = "Routes_to_run (comma or 'all')"; ws["B10"] = "all"
    ws["A11"] = "Speeds_to_run (comma or 'all')"; ws["B11"] = "all"
    bold = Font(bold=True)
    for r in range(1, 12):
        ws[f"A{r}"].font = bold
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "A2"
    return ws

def _parse_list_or_all(value, kind="int"):
    """
    "all" または空欄 → None
    それ以外 → カンマ/空白区切りのリストにして返す
    kind="int" or "float"
    """
    if value is None:
        return None
    s = str(value).strip().lower()
    if s == "" or s == "all":
        return None
    parts = [p for p in re.split(r"[\s,]+", s) if p]
    if kind == "int":
        try:
            return [int(p) for p in parts]
        except ValueError as e:
            raise ValueError(f'Routes_to_run は整数のカンマ区切りか "all" で指定してください。いま: {value!r}') from e
    else:
        try:
            return [float(p) for p in parts]
        except ValueError as e:
            raise ValueError(f'Speeds_to_run は数値のカンマ区切りか "all" で指定してください。いま: {value!r}') from e

def read_config(cfg):
    def fcell(addr, name):
        v = cfg[addr].value
        if v is None or str(v).strip() == "":
            raise ValueError(f'CONFIG の {name} ({addr}) が空です。')
        return float(v)

    C = fcell("B1", "Cycle_C_sec")
    step = fcell("B2", "Step (cycle fraction)")
    P = fcell("B3", "P")
    sat = fcell("B4", "sat_flow_veh_per_s")

    def opt_float(addr):
        v = cfg[addr].value
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        return float(s)

    g0 = opt_float("B5")
    g1 = fcell("B6", "g1")
    g2 = fcell("B7", "g2")
    g3 = opt_float("B8")
    if g0 is None:
        g0 = P
    if g3 is None:
        g3 = P
    g = [g0, g1, g2, g3]

    routes = _parse_list_or_all(cfg["B10"].value, kind="int")
    speeds = _parse_list_or_all(cfg["B11"].value, kind="float")

    # validation
    if not (C > 0):
        raise ValueError("Cycle_C_sec は正の値にしてください。")
    if not (0 < step <= 1):
        raise ValueError("Step は (0,1] の範囲にしてください。例: 0.01")
    T = 1 / step
    if abs(T - round(T)) > 1e-9:
        raise ValueError("Step は 1/Step が整数になる値にしてください。例: 0.01, 0.02, 0.005")
    if not (0 < P <= 1):
        raise ValueError("P は (0,1] の範囲にしてください。")
    if not (sat > 0):
        raise ValueError("sat_flow_veh_per_s は正の値にしてください。")
    for i, gi in enumerate(g):
        if not (0 < gi <= 1):
            raise ValueError(f"g{i} は (0,1] の範囲にしてください。いま g{i}={gi}")
    return C, step, P, sat, g, routes, speeds

def iter_rows(ws, col_route, col_speed):
    cur = None
    for r in range(2, ws.max_row + 1):
        rn = ws.cell(r, col_route).value
        if rn is not None:
            cur = rn
        sp = ws.cell(r, col_speed).value
        if sp is None:
            continue
        yield cur, float(sp), r

def get_tau_from_row(ws, r, col_speed, col_L01, col_L12, col_L23):
    v = float(ws.cell(r, col_speed).value)  # km/h
    mps = v * 1000.0 / 3600.0
    L01 = float(ws.cell(r, col_L01).value)  # m
    L12 = float(ws.cell(r, col_L12).value)
    L23 = float(ws.cell(r, col_L23).value)
    return (L01 / mps, L12 / mps, L23 / mps)

def simulate_one_dir(T, dt, sat, P, green_masks, delay_steps, origin_node):
    d01, d12, d23 = delay_steps
    maxd = max(d01, d12, d23)

    arr01 = [0.0] * (T + maxd + 5)
    arr12 = [0.0] * (T + maxd + 5)
    arr23 = [0.0] * (T + maxd + 5)

    q = [0.0, 0.0, 0.0, 0.0]
    delay = [0.0, 0.0, 0.0, 0.0]
    vehicles_entered = 0.0

    mask0 = green_masks[origin_node]
    try:
        start_idx = mask0.index(True)
    except ValueError:
        start_idx = 0
    dur_idx = int(round(P * T))
    demand = build_mask(T, start_idx, dur_idx)

    for t in range(T):
        if demand[t]:
            a = sat * dt
            q[origin_node] += a
            vehicles_entered += a

        if origin_node == 0:
            q[1] += arr01[t]
            q[2] += arr12[t]
            q[3] += arr23[t]
            order = [0, 1, 2, 3]
        else:
            q[2] += arr23[t]
            q[1] += arr12[t]
            q[0] += arr01[t]
            order = [3, 2, 1, 0]

        for i in range(4):
            delay[i] += q[i] * dt

        for i in order:
            cap = sat * dt if green_masks[i][t] else 0.0
            dep = q[i] if q[i] < cap else cap
            q[i] -= dep

            if origin_node == 0:
                if i == 0:
                    arr01[t + d01] += dep
                elif i == 1:
                    arr12[t + d12] += dep
                elif i == 2:
                    arr23[t + d23] += dep
            else:
                if i == 3:
                    arr23[t + d23] += dep
                elif i == 2:
                    arr12[t + d12] += dep
                elif i == 1:
                    arr01[t + d01] += dep

    return delay, vehicles_entered

def eval_offsets_pointqueue(x1, x2, x3, tau01, tau12, tau23, C, step, P, sat, g):
    dt = step * C
    T = int(round(1 / step))

    d01 = int(round(tau01 / dt))
    d12 = int(round(tau12 / dt))
    d23 = int(round(tau23 / dt))

    xs = [0.0, x1, x2, x3]
    green_masks = []
    for i in range(4):
        start = int(round(xs[i] * T)) % T
        dur = int(round(g[i] * T))
        green_masks.append(build_mask(T, start, dur))

    delay_f, veh_f = simulate_one_dir(T, dt, sat, P, green_masks, (d01, d12, d23), origin_node=0)
    delay_b, veh_b = simulate_one_dir(T, dt, sat, P, green_masks, (d01, d12, d23), origin_node=3)

    delay_total = [delay_f[i] + delay_b[i] for i in range(4)]
    veh_total = veh_f + veh_b
    if veh_total <= 0:
        return [float("inf")] * 4, float("inf")

    d_nodes = [delay_total[i] / veh_total for i in range(4)]
    return d_nodes, sum(d_nodes)

def topk_offsets(tau01, tau12, tau23, C, step, P, sat, g, k=10):
    n = int(round(1 / step))
    grid = [round(i * step, 10) for i in range(n)]

    heap = []
    for x1 in grid:
        for x2 in grid:
            for x3 in grid:
                d_nodes, d_total = eval_offsets_pointqueue(x1, x2, x3, tau01, tau12, tau23, C, step, P, sat, g)
                item = (-d_total, x1, x2, x3, d_nodes[0], d_nodes[1], d_nodes[2], d_nodes[3])
                if len(heap) < k:
                    heapq.heappush(heap, item)
                else:
                    if item[0] > heap[0][0]:
                        heapq.heapreplace(heap, item)

    res = [(-h[0], h[1], h[2], h[3], h[4], h[5], h[6], h[7]) for h in heap]
    res.sort(key=lambda x: x[0])
    return res

def write_best_to_main(ws, r, col, best):
    dtotal, x1, x2, x3, d0, d1, d2, d3 = best
    ws.cell(r, col["x0opt"]).value = 0.0
    ws.cell(r, col["x1opt"]).value = x1
    ws.cell(r, col["x2opt"]).value = x2
    ws.cell(r, col["x3opt"]).value = x3
    ws.cell(r, col["d0opt(s/veh)"]).value = d0
    ws.cell(r, col["d1opt(s/veh)"]).value = d1
    ws.cell(r, col["d2opt(s/veh)"]).value = d2
    ws.cell(r, col["d3opt(s/veh)"]).value = d3
    ws.cell(r, col["dopt(s/veh)"]).value = dtotal

def make_detail_sheet(wb, route, speed, top10, C, step, P, sat, g):
    title = sanitize_sheet_name(f"R{route}_S{speed:g}")
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws = wb.create_sheet(title)

    ws["A1"] = "route"; ws["B1"] = route
    ws["A2"] = "speed(km/h)"; ws["B2"] = speed
    ws["A3"] = "C(sec)"; ws["B3"] = C
    ws["A4"] = "step"; ws["B4"] = step
    ws["A5"] = "P"; ws["B5"] = P
    ws["A6"] = "sat_flow"; ws["B6"] = sat
    ws["A7"] = "g0,g1,g2,g3"; ws["B7"] = str(g)

    ws["A9"] = "rank"
    headers = ["d_total(s/veh)", "x1", "x2", "x3", "d0", "d1", "d2", "d3"]
    for j, h in enumerate(headers, start=2):
        ws.cell(9, j).value = h
    bold = Font(bold=True)
    for c in range(1, 10):
        ws.cell(9, c).font = bold

    for i, row in enumerate(top10, start=1):
        dtotal, x1, x2, x3, d0, d1, d2, d3 = row
        ws.cell(9 + i, 1).value = i
        for j, v in enumerate([dtotal, x1, x2, x3, d0, d1, d2, d3], start=2):
            ws.cell(9 + i, j).value = float(v)

    ws.freeze_panes = "A10"
    ws.column_dimensions["A"].width = 6
    for c in "BCDEFGHI":
        ws.column_dimensions[c].width = 16
    return title

def make_summary_sheet(wb, rows):
    if SUMMARY_SHEET in wb.sheetnames:
        wb.remove(wb[SUMMARY_SHEET])
    ws = wb.create_sheet(SUMMARY_SHEET, 0)
    headers = ["route", "speed(km/h)", "rank", "d_total(s/veh)", "x1", "x2", "x3", "d0", "d1", "d2", "d3", "detail_sheet"]
    for j, h in enumerate(headers, start=1):
        ws.cell(1, j).value = h
        ws.cell(1, j).font = Font(bold=True)
    for i, rr in enumerate(rows, start=2):
        for j, v in enumerate(rr, start=1):
            ws.cell(i, j).value = v
    ws.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.column_dimensions["L"].width = 22
    return ws

def main(xlsx_path: str, out_path: str | None = None, sheet_name: str | None = None):
    wb = openpyxl.load_workbook(xlsx_path)
    cfg = ensure_config_sheet(wb)
    C, step, P, sat, g, target_routes, target_speeds = read_config(cfg)

    ws = find_data_sheet(wb, preferred_name=sheet_name)
    col = header_map(ws)
    missing = [h for h in REQUIRED_HEADERS if h not in col]
    if missing:
        raise ValueError(f'入力シート "{ws.title}" に必要な見出しがありません: {missing}')

    summary_rows = []

    for route, speed, r in iter_rows(ws, col["路線番号"], col["系統速度"]):
        if route is None:
            continue
        # フィルタ（Noneならall）
        if target_routes is not None and int(route) not in target_routes:
            continue
        if target_speeds is not None and float(speed) not in target_speeds:
            continue

        tau01, tau12, tau23 = get_tau_from_row(ws, r, col["系統速度"], col["link(0,1)"], col["link(1,2)"], col["link(2,3)"])
        top10 = topk_offsets(tau01, tau12, tau23, C, step, P, sat, g, k=10)

        write_best_to_main(ws, r, col, top10[0])
        detail = make_detail_sheet(wb, int(route), speed, top10, C, step, P, sat, g)

        for rank, row in enumerate(top10, start=1):
            dtotal, x1, x2, x3, d0, d1, d2, d3 = row
            summary_rows.append([int(route), float(speed), rank, dtotal, x1, x2, x3, d0, d1, d2, d3, detail])

        print(f"done: route={int(route)} speed={speed:g} best_d={top10[0][0]:.6g}")

    make_summary_sheet(wb, summary_rows)

    if out_path is None:
        p = Path(xlsx_path)
        out_path = str(p.with_name(p.stem + "_out.xlsx"))

    wb.save(out_path)
    print(f"Saved: {out_path}")

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="input xlsx (e.g. experiment.xlsx)")
    ap.add_argument("--out", default=None, help="output xlsx (default: input_stem_out.xlsx)")
    ap.add_argument("--sheet", default=None, help="data sheet name (optional). If omitted, auto-detect.")
    args = ap.parse_args()
    main(args.xlsx, args.out, args.sheet)
