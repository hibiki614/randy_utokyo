# -*- coding: utf-8 -*-
"""
Created on Sat Oct 18 03:34:34 2025

@author: OguchiLab
"""

import pandas as pd
import matplotlib.pyplot as plt
import math
from datetime import datetime, timedelta
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

cycle = 120
vmax = 50 / 3.6
print('\nCurrent setting for cycle length = '+str(cycle)+' seconds')
print('Current setting for maximum speed = '+str(np.round(vmax*3.6, 2))+' km/h')
print('\n1-100 : link lengths between 80 and 200 metres')
print('101-200 : link lengths between 100 and 400 metres')
print('201-300 : link lengths between 200 and 600 metres')
road = int(input('\nChoice (integer beween 1 and 300) = '))


wb = load_workbook(filename = "data.xlsx", read_only = True, data_only = True)



intersections = dict([(k, dict(green_split = float(wb['greens'][get_column_letter(k + 1) + str(road)].value) / cycle)) for k in range(15)])
links = dict([(k, dict(length = int(wb['lengths'][get_column_letter(k + 1) + str(road)].value))) for k in range(14)])

P = intersections[0]['green_split']
intersections[0]['position'] = 0
intersections[0]['x'] = 0
start1l = 0
end1r = 0
for k in range(14) :
    link = links[k]
    intersections[k+1]['position'] = intersections[k]['position'] + link['length']
    green_split = intersections[k+1]['green_split']
    length = link['length']
    if end1r - start1l == 0 :
        link['first_upper_bound'] = 1e6
    else :
        link['first_upper_bound'] = 2 * length / cycle / (end1r - start1l)
    if green_split + end1r - start1l - P == 0 :
        link['first_lower_bound'] = 1e6
    else :
        link['first_lower_bound'] = 2 * length / cycle / (green_split + end1r - start1l - P)
    if 1 + end1r - start1l == 0 :
        link['second_upper_bound'] = 1e6
    else :
        link['second_upper_bound'] = 2 * length / cycle / (1 + end1r - start1l)
    if 1 + green_split + end1r - start1l - P == 0 :
        link['second_lower_bound'] = 1e6
    else :
        link['second_lower_bound'] = 2 * length / cycle / (1 + green_split + end1r - start1l - P)
    if 2 + end1r - start1l == 0 :
        link['third_upper_bound'] = 1e6
    else :
        link['third_upper_bound'] = 2 * length / cycle / (2 + end1r - start1l)
    if 2 + green_split + end1r - start1l - P == 0 :
        link['third_lower_bound'] = 1e6
    else :
        link['third_lower_bound'] = 2 * length / cycle / (2 + green_split + end1r - start1l - P)
    if 3 + end1r - start1l == 0 :
        link['fourth_upper_bound'] = 1e6
    else :
        link['fourth_upper_bound'] = 2 * length / cycle / (3 + end1r - start1l)
    if 3 + green_split + end1r - start1l - P == 0 :
        link['fourth_lower_bound'] = 1e6
    else :
        link['fourth_lower_bound'] = 2 * length / cycle / (3 + green_split + end1r - start1l - P)
    """if end1r - start1l == 0 :
        link['fifth_upper_bound'] = 1e6
    else :
        link['fifth_upper_bound'] = 2 * length / cycle / (end1r - start1l)
    if green_split + end1r - start1l - P == 0 :
        link['fifth_lower_bound'] = 1e6
    else :
        link['fifth_lower_bound'] = 2 * length / cycle / (green_split + end1r - start1l - P)
    if 1 + end1r - start1l == 0 :
        link['sixth_upper_bound'] = 1e6
    else :
        link['sixth_upper_bound'] = 2 * length / cycle / (1 + end1r - start1l)
    if 1 + green_split + end1r - start1l - P == 0 :
        link['sixth_lower_bound'] = 1e6
    else :
        link['sixth_lower_bound'] = 2 * length / cycle / (1 + green_split + end1r - start1l - P)
    if 2 + end1r - start1l == 0 :
        link['seventh_upper_bound'] = 1e6
    else :
        link['seventh_upper_bound'] = 2 * length / cycle / (2 + end1r - start1l)
    if 2 + green_split + end1r - start1l - P == 0 :
        link['seventh_lower_bound'] = 1e6
    else :
        link['seventh_lower_bound'] = 2 * length / cycle / (2 + green_split + end1r - start1l - P)
    if 3 + end1r - start1l == 0 :
        link['eighth_upper_bound'] = 1e6
    else :
        link['eighth_upper_bound'] = 2 * length / cycle / (3 + end1r - start1l)
    if 3 + green_split + end1r - start1l - P == 0 :
        link['eighth_lower_bound'] = 1e6
    else :
        link['eighth_lower_bound'] = 2 * length / cycle / (3 + green_split + end1r - start1l - P)"""
    if (link['first_upper_bound'] >= link['first_lower_bound']) and (link['first_upper_bound'] >= vmax) and (vmax >= link['first_lower_bound']) :
        link['v50prox'] = vmax
        choice = True
    elif (link['second_upper_bound'] >= link['second_lower_bound']) and (link['second_upper_bound'] >= vmax) and (vmax >= link['second_lower_bound']) :
        link['v50prox'] = vmax
        choice = True
    elif (link['third_upper_bound'] >= link['third_lower_bound']) and (link['third_upper_bound'] >= vmax) and (vmax >= link['third_lower_bound']) :
        link['v50prox'] = vmax
        choice = True
    elif (link['fourth_upper_bound'] >= link['fourth_lower_bound']) and (link['fourth_upper_bound'] >= vmax) and (vmax >= link['fourth_lower_bound']) :
        link['v50prox'] = vmax
        choice = True
    else :
        speeds = [link['first_upper_bound'], link['second_upper_bound'], link['third_upper_bound'], link['fourth_upper_bound']]#, link['fifth_upper_bound'], link['sixth_upper_bound'], link['seventh_upper_bound'], link['eighth_upper_bound']]
        link['v50prox'] = max([v for v in speeds if v <= vmax and v > 0])
        choice = speeds.index(link['v50prox']) < 4
    """elif (link['fifth_upper_bound'] >= link['fifth_lower_bound']) and (link['fifth_upper_bound'] >= vmax) and (vmax >= link['fifth_lower_bound']) :
        link['v50prox'] = vmax
        choice = False
    elif (link['sixth_upper_bound'] >= link['sixth_lower_bound']) and (link['sixth_upper_bound'] >= vmax) and (vmax >= link['sixth_lower_bound']) :
        link['v50prox'] = vmax
        choice = False
    elif (link['seventh_upper_bound'] >= link['seventh_lower_bound']) and (link['seventh_upper_bound'] >= vmax) and (vmax >= link['seventh_lower_bound']) :
        link['v50prox'] = vmax
        choice = False
    elif (link['eighth_upper_bound'] >= link['eighth_lower_bound']) and (link['eighth_upper_bound'] >= vmax) and (vmax >= link['eighth_lower_bound']) :
        link['v50prox'] = vmax
        choice = False"""
    dt = length / link['v50prox']
    link['dt50prox'] = dt
    start1l = (start1l + dt / cycle) % 1
    end1r = (end1r - dt / cycle) % 1
    intersections[k+1]['x'] = end1r * choice + start1l * (1 - choice)


print('\n'.join([str(links[k]['v50prox']) for k in range(14)]))
#exit()
left_stream_dts = [cycle * intersections[0]['x']] + [links[k]['dt50prox'] for k in range(14)]
left_stream_first = [sum(left_stream_dts[:k]) % cycle for k in range(1,16)]
left_stream_last = [(cycle * intersections[0]['green_split'] + sum(left_stream_dts[:k])) % cycle for k in range(1,16)]
right_stream_dts = [cycle * (1 + intersections[0]['x'])] + [- links[k]['dt50prox'] for k in range(14)]
right_stream_first = [sum(right_stream_dts[:k]) % cycle for k in range(1,16)]
right_stream_last = [(cycle * intersections[0]['green_split'] + sum(right_stream_dts[:k])) % cycle for k in range(1,16)]
nC=3
fig, ax = plt.subplots(figsize=(16, 9))
ax.grid(':',alpha=0.5)
for k in range(15) :
    intersection = intersections[k]
    ax.plot([intersection['position']] * 2, [[cycle * (intersection['x'] + i) for i in range(-2, nC + 1)] , [cycle * (intersection['x'] + intersection['green_split'] + i) for i in range(-2, nC + 1)]], 'g:')
    ax.plot([intersection['position']] * 2, [[cycle * (intersection['x'] + intersection['green_split'] + i) for i in range(-2, nC + 1)] , [cycle * (intersection['x'] + i + 1) for i in range(-2, nC + 1)]], 'r-')
    if k < 14 :
        next_intersection = intersections[k+1]
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + left_stream_first[k] for i in range(nC)] , [cycle * i + left_stream_first[k] + links[k]['dt50prox'] for i in range(nC)]], 'k')
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + left_stream_last[k] for i in range(nC)] , [cycle * i + left_stream_last[k] + links[k]['dt50prox'] for i in range(nC)]], 'k', alpha=.5)
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + right_stream_first[k] for i in range(nC)] , [cycle * i + right_stream_first[k] - links[k]['dt50prox'] for i in range(nC)]], 'b')
        ax.plot([intersection['position'], next_intersection['position']], [[cycle * i + right_stream_last[k] for i in range(nC)] , [cycle * i + right_stream_last[k] - links[k]['dt50prox'] for i in range(nC)]], 'b', alpha=.5)
ax.set_xlim([- 50, intersections[14]['position'] + 50])
ax.set_ylim([- cycle, (nC + 1 ) * cycle])
ax.set_xlabel('Space (m)')
ax.set_ylabel('Time (s)')
ax.set_xticks([intersections[k]['position'] for k in range(15)])
ax.set_yticks([cycle * i for i in range(nC+1)])
ax.set_title('Travel time = '+str(int(np.round(sum([links[k]['dt50prox'] for k in range(14)]))))+' s [best = '+str(int(np.round(intersections[14]['position'] / vmax)))+' s, worst = '+str(int(np.round(intersections[14]['position'] / vmax + cycle * sum([1 - intersections[k]['green_split'] for k in range(14)]))))+' s] | Average speed = '+str(np.round(intersections[14]['position'] / sum([links[k]['dt50prox'] for k in range(14)]) * 3.6, 2))+' km/h [best = '+str(np.round(vmax * 3.6, 2))+' km/h, worst = '+str(np.round(intersections[14]['position'] / (intersections[14]['position'] / vmax + cycle * sum([1 - intersections[k]['green_split'] for k in range(14)])) * 3.6, 2))+' km/h]')
plt.tight_layout()
plt.show()