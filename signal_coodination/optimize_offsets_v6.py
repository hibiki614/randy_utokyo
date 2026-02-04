#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
optimize_offsets_v6.py

v6: ユーザの前提を明確に反映した版
- x は cycle fraction（0〜1）。探索は x1,x2,x3 を 0.01 刻みで総当たり（x0=0固定）
- 端点(0,3)の需要は「その交差点の青開始と同時に出発し始める」
    * 端点でキューを作らず（=端点遅れ0）、青の間だけ sat_flow でリンクへ直接注入
    * 需要窓は各端点の青開始から P*C 秒間
        - 0端: t_start0 = x0*C (=0), t_end0 = t_start0 + P*C
        - 3端: t_start3 = x3*C,        t_end3 = t_start3 + P*C
- sat_flow_veh_per_s は
    * 需要（端点からの注入率）
    * 青中の放出率（交差点1,2でのサービス率）
  の両方に使用
- forward/backward は別車線（容量共有なし）：交差点1,2で各方向がそれぞれ sat で放出できる
- 無拡散：リンクは固定走行時間だけ遅れる delay line（dt刻みで到着）
- 評価：両端から注入された車群が「最後尾までゴールしきるまで」発生した停止時間合計（∫Q dt）を
        総注入台数で割って s/veh
    * 端点(0,3)は遅れ0として扱い、遅れ積分は交差点1,2のみ（d0=d3=0）

Excel I/O
- CONFIG シートから C, step, P, sat, g0..g3, 実行対象の路線/速度を読み込み
- データシートは必要な見出しが揃う最初のシートを自動検出（--sheet で指定可）
- 出力：元シートに最良解を書込み、Summary_Top10 と各ケース詳細シートを作成

実行例
  python optimize_offsets_v6.py experiment.xlsx --out experiment_out.xlsx
"""

import re
import heapq
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

CONFIG_SHEET = "CONFIG"
SUMMARY_SHEET = "Summary_Top10"

REQUIRED_HEADERS = [
    "路線番号", "系統速度",
    "link(0,1)", "link(1,2)", "link(2,3)",
    "x0opt", "x1opt", "x2opt", "x3opt",
    "d0opt(s/veh)", "d1opt(s/veh)", "d2opt(s/veh)", "d3opt(s/veh)", "dopt(s/veh)"
]

def sanitize_sheet_name(name: str) -> str:
    name = re.sub(r'[\[\]\*\?/\\:]', '_', name)
    return name[:31]

def header_map(ws):
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    return {h: i + 1 for i, h in enumerate(headers) if h is not None}

def find_data_sheet(wb, preferred_name=None):
    if preferred_name:
        if preferred_name not in wb.sheetnames:
            raise ValueError(f'--sheet で指定した "{preferred_name}" が見つかりません。')
        ws = wb[preferred_name]
        col = header_map(ws)
        missing = [h for h in REQUIRED_HEADERS if h not in col]
        if missing:
            raise ValueError(f'"{preferred_name}" に必要な見出しがありません: {missing}')
        return ws

    for name in wb.sheetnames:
        ws = wb[name]
        col = header_map(ws)
        missing = [h for h in REQUIRED_HEADERS if h not in col]
        if not missing:
            return ws
    raise ValueError(f"必要な見出し {REQUIRED_HEADERS} をすべて含むシートが見つかりません。")

def ensure_config_sheet(wb):
    if CONFIG_SHEET in wb.sheetnames:
        return wb[CONFIG_SHEET]
    ws = wb.create_sheet(CONFIG_SHEET, 0)
    ws["A1"]  = "Cycle_C_sec"; ws["B1"]  = 120
    ws["A2"]  = "Step (cycle fraction)"; ws["B2"]  = 0.01
    ws["A3"]  = "P (demand time rate, same both dirs)"; ws["B3"]  = 0.4
    ws["A4"]  = "sat_flow_veh_per_s"; ws["B4"]  = 0.5
    ws["A5"]  = "g0 (if blank, use P)"; ws["B5"]  = ""
    ws["A6"]  = "g1"; ws["B6"]  = 0.6
    ws["A7"]  = "g2"; ws["B7"]  = 0.6
    ws["A8"]  = "g3 (if blank, use P)"; ws["B8"]  = ""
    ws["A10"] = "Routes_to_run (comma or 'all')"; ws["B10"] = "all"
    ws["A11"] = "Speeds_to_run (comma or 'all')"; ws["B11"] = "all"
    bold = Font(bold=True)
    for r in range(1, 12):
        ws[f"A{r}"].font = bold
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "A2"
    return ws

def _parse_list_or_all(value, kind="int"):
    if value is None:
        return None
    s = str(value).strip().lower()
    if s == "" or s == "all":
        return None
    parts = [p for p in re.split(r"[\s,]+", s) if p]
    if kind == "int":
        try:
            return [int(p) for p in parts]
        except ValueError as e:
            raise ValueError(f'Routes_to_run は整数のカンマ区切りか "all" で指定してください。いま: {value!r}') from e
    else:
        try:
            return [float(p) for p in parts]
        except ValueError as e:
            raise ValueError(f'Speeds_to_run は数値のカンマ区切りか "all" で指定してください。いま: {value!r}') from e

def read_config(cfg):
    def fcell(addr, name):
        v = cfg[addr].value
        if v is None or str(v).strip() == "":
            raise ValueError(f'CONFIG の {name} ({addr}) が空です。')
        return float(v)

    C = fcell("B1", "Cycle_C_sec")
    step = fcell("B2", "Step (cycle fraction)")
    P = fcell("B3", "P")
    sat = fcell("B4", "sat_flow_veh_per_s")

    def opt_float(addr):
        v = cfg[addr].value
        if v is None:
            return None
        s = str(v).strip()
        if s == "":
            return None
        return float(s)

    g0 = opt_float("B5")
    g1 = fcell("B6", "g1")
    g2 = fcell("B7", "g2")
    g3 = opt_float("B8")
    if g0 is None: g0 = P
    if g3 is None: g3 = P
    g = [g0, g1, g2, g3]

    routes = _parse_list_or_all(cfg["B9"].value, kind="int")
    speeds = _parse_list_or_all(cfg["B10"].value, kind="float")

    if not (C > 0):
        raise ValueError("Cycle_C_sec は正の値にしてください。")
    if not (0 < step <= 1):
        raise ValueError("Step は (0,1] の範囲にしてください。例: 0.01")
    inv = 1 / step
    if abs(inv - round(inv)) > 1e-9:
        raise ValueError("Step は 1/Step が整数になる値にしてください。例: 0.01, 0.02, 0.005")
    if not (0 < P <= 1):
        raise ValueError("P は (0,1] の範囲にしてください。")
    if not (sat > 0):
        raise ValueError("sat_flow_veh_per_s は正の値にしてください。")
    for i, gi in enumerate(g):
        if not (0 < gi <= 1):
            raise ValueError(f"g{i} は (0,1] の範囲にしてください。いま g{i}={gi}")
    return C, step, P, sat, g, routes, speeds

def iter_rows(ws, col_route, col_speed):
    cur = None
    for r in range(2, ws.max_row + 1):
        rn = ws.cell(r, col_route).value
        if rn is not None:
            cur = rn
        sp = ws.cell(r, col_speed).value
        if sp is None:
            continue
        yield cur, float(sp), r

def get_tau_from_row(ws, r, col_speed, col_L01, col_L12, col_L23):
    v = float(ws.cell(r, col_speed).value)  # km/h
    mps = v * 1000.0 / 3600.0
    L01 = float(ws.cell(r, col_L01).value)  # m
    L12 = float(ws.cell(r, col_L12).value)
    L23 = float(ws.cell(r, col_L23).value)
    return (L01 / mps, L12 / mps, L23 / mps)

def is_green(t: float, C: float, green_start: float, green_dur: float) -> bool:
    if green_dur <= 0:
        return False
    x = (t - green_start) % C
    return x < green_dur

def _make_pipe(delay_s: float, dt: float):
    n = max(0, int(round(delay_s / dt)))
    return [0.0] * (n + 1)

def simulate_bidir_until_finish(C, step, P, sat, xs, g, tau01, tau12, tau23):
    """
    v6 の評価シミュレーション
    - 端点(0,3)はキューを作らず、青の間だけ sat でリンクへ直接注入（端点遅れ=0）
    - 交差点1,2でのキュー停止を積分して遅れ（veh*s）を得る
    - 需要窓は各端点の青開始から P*C 秒間（0端は x0=0固定、3端は x3 で動く）
    """
    dt = step * C
    demand_T = P * C

    # Terminal demand windows (start at each terminal green start)
    t_start0 = xs[0] * C          # x0 fixed 0
    t_end0   = t_start0 + demand_T
    t_start3 = xs[3] * C
    t_end3   = t_start3 + demand_T
    t_end_all = max(t_end0, t_end3)

    # Total injected vehicles (by design we assume demand window defines amount)
    vehicles_total = 2.0 * sat * demand_T

    green_start = [xs[i] * C for i in range(4)]
    green_dur   = [g[i]  * C for i in range(4)]

    # delay lines (forward)
    pipe01_f = _make_pipe(tau01, dt)  # from terminal 0 (injection goes here)
    pipe12_f = _make_pipe(tau12, dt)
    pipe23_f = _make_pipe(tau23, dt)  # to terminal 3 (sink)

    # delay lines (backward)
    pipe23_b = _make_pipe(tau23, dt)  # from terminal 3 (injection goes here)
    pipe12_b = _make_pipe(tau12, dt)
    pipe01_b = _make_pipe(tau01, dt)  # to terminal 0 (sink)

    # queues only at nodes 1 and 2 (terminals excluded)
    q1_f = 0.0; q2_f = 0.0
    q2_b = 0.0; q1_b = 0.0

    # delay accumulators per node (veh*s): d0=d3=0 by construction
    delay = [0.0, 0.0, 0.0, 0.0]

    # safety steps
    max_steps = int(round((t_end_all + 10*C + (tau01+tau12+tau23)) / dt)) + 4000

    t = 0.0
    cap = sat * dt

    for _ in range(max_steps):
        # 1) arrivals from links to nodes
        q1_f += pipe01_f[0]
        q2_f += pipe12_f[0]
        # pipe23_f[0] arrives at sink (node3): ignore queue/delay

        q2_b += pipe23_b[0]
        q1_b += pipe12_b[0]
        # pipe01_b[0] arrives at sink (node0): ignore queue/delay

        # shift pipes
        pipe01_f = pipe01_f[1:] + [0.0]
        pipe12_f = pipe12_f[1:] + [0.0]
        pipe23_f = pipe23_f[1:] + [0.0]

        pipe23_b = pipe23_b[1:] + [0.0]
        pipe12_b = pipe12_b[1:] + [0.0]
        pipe01_b = pipe01_b[1:] + [0.0]

        # 2) terminal injections directly into first links when terminal is green AND within its demand window
        # forward injection at node0 into link(0,1)
        if (t_start0 <= t < t_end0) and is_green(t, C, green_start[0], green_dur[0]):
            pipe01_f[-1] += cap
        # backward injection at node3 into link(2,3) backward (3->2)
        if (t_start3 <= t < t_end3) and is_green(t, C, green_start[3], green_dur[3]):
            pipe23_b[-1] += cap

        # 3) accumulate delay only at nodes 1 and 2 (both directions)
        delay[1] += (q1_f + q1_b) * dt
        delay[2] += (q2_f + q2_b) * dt
        # delay[0]=delay[3]=0

        # 4) discharge at nodes 1 and 2 (per direction, no sharing)
        # node1
        if is_green(t, C, green_start[1], green_dur[1]):
            dep1_f = cap if q1_f >= cap else q1_f
            dep1_b = cap if q1_b >= cap else q1_b
            q1_f -= dep1_f; q1_b -= dep1_b
            pipe12_f[-1] += dep1_f
            pipe01_b[-1] += dep1_b  # node1 -> node0 (backward last link)
        # node2
        if is_green(t, C, green_start[2], green_dur[2]):
            dep2_f = cap if q2_f >= cap else q2_f
            dep2_b = cap if q2_b >= cap else q2_b
            q2_f -= dep2_f; q2_b -= dep2_b
            pipe23_f[-1] += dep2_f  # node2 -> node3 (forward last link)
            pipe12_b[-1] += dep2_b  # node2 -> node1 (backward middle link)

        # 5) finish condition after both demand windows ended:
        # all internal queues empty + all pipes empty (sinks don't queue)
        if t >= t_end_all:
            if (q1_f + q2_f + q1_b + q2_b) == 0.0:
                if (sum(pipe01_f)+sum(pipe12_f)+sum(pipe23_f)+sum(pipe23_b)+sum(pipe12_b)+sum(pipe01_b)) == 0.0:
                    break

        t += dt

    return delay, vehicles_total

def eval_offsets_pointqueue(x1, x2, x3, tau01, tau12, tau23, C, step, P, sat, g):
    xs = [0.0, x1, x2, x3]
    delay_veh_s, veh_total = simulate_bidir_until_finish(C, step, P, sat, xs, g, tau01, tau12, tau23)
    if veh_total <= 0:
        return [float("inf")] * 4, float("inf")
    d_nodes = [delay_veh_s[i] / veh_total for i in range(4)]  # s/veh
    return d_nodes, sum(d_nodes)

def topk_offsets(tau01, tau12, tau23, C, step, P, sat, g, k=10, progress_every=200000):
    n = int(round(1 / step))
    grid = [round(i * step, 10) for i in range(n)]
    total_iters = n * n * n

    heap = []
    cnt = 0

    for x1 in grid:
        for x2 in grid:
            for x3 in grid:
                cnt += 1
                if progress_every and (cnt == 1 or cnt % progress_every == 0 or cnt == total_iters):
                    pct = 100.0 * cnt / total_iters
                    print(f"    search: {cnt:,}/{total_iters:,} ({pct:.1f}%)", flush=True)

                d_nodes, d_total = eval_offsets_pointqueue(x1, x2, x3, tau01, tau12, tau23, C, step, P, sat, g)
                item = (-d_total, x1, x2, x3, d_nodes[0], d_nodes[1], d_nodes[2], d_nodes[3])
                if len(heap) < k:
                    heapq.heappush(heap, item)
                else:
                    if item[0] > heap[0][0]:
                        heapq.heapreplace(heap, item)

    res = [(-h[0], h[1], h[2], h[3], h[4], h[5], h[6], h[7]) for h in heap]
    res.sort(key=lambda x: x[0])
    return res

def write_best_to_main(ws, r, col, best):
    dtotal, x1, x2, x3, d0, d1, d2, d3 = best
    ws.cell(r, col["x0opt"]).value = 0.0
    ws.cell(r, col["x1opt"]).value = x1
    ws.cell(r, col["x2opt"]).value = x2
    ws.cell(r, col["x3opt"]).value = x3
    ws.cell(r, col["d0opt(s/veh)"]).value = d0
    ws.cell(r, col["d1opt(s/veh)"]).value = d1
    ws.cell(r, col["d2opt(s/veh)"]).value = d2
    ws.cell(r, col["d3opt(s/veh)"]).value = d3
    ws.cell(r, col["dopt(s/veh)"]).value = dtotal

def make_detail_sheet(wb, route, speed, top10, C, step, P, sat, g):
    title = sanitize_sheet_name(f"R{route}_S{speed:g}")
    if title in wb.sheetnames:
        wb.remove(wb[title])
    ws = wb.create_sheet(title)

    ws["A1"] = "route"; ws["B1"] = route
    ws["A2"] = "speed(km/h)"; ws["B2"] = speed
    ws["A3"] = "C(sec)"; ws["B3"] = C
    ws["A4"] = "step"; ws["B4"] = step
    ws["A5"] = "P"; ws["B5"] = P
    ws["A6"] = "sat_flow"; ws["B6"] = sat
    ws["A7"] = "g0,g1,g2,g3"; ws["B7"] = str(g)

    ws["A9"] = "rank"
    headers = ["d_total(s/veh)", "x1", "x2", "x3", "d0", "d1", "d2", "d3"]
    for j, h in enumerate(headers, start=2):
        ws.cell(9, j).value = h
    bold = Font(bold=True)
    for c in range(1, 10):
        ws.cell(9, c).font = bold

    for i, row in enumerate(top10, start=1):
        dtotal, x1, x2, x3, d0, d1, d2, d3 = row
        ws.cell(9 + i, 1).value = i
        for j, v in enumerate([dtotal, x1, x2, x3, d0, d1, d2, d3], start=2):
            ws.cell(9 + i, j).value = float(v)

    ws.freeze_panes = "A10"
    ws.column_dimensions["A"].width = 6
    for c in "BCDEFGHI":
        ws.column_dimensions[c].width = 16
    return title

def make_summary_sheet(wb, rows):
    if SUMMARY_SHEET in wb.sheetnames:
        wb.remove(wb[SUMMARY_SHEET])
    ws = wb.create_sheet(SUMMARY_SHEET, 0)
    headers = ["route", "speed(km/h)", "rank", "d_total(s/veh)", "x1", "x2", "x3", "d0", "d1", "d2", "d3", "detail_sheet"]
    for j, h in enumerate(headers, start=1):
        ws.cell(1, j).value = h
        ws.cell(1, j).font = Font(bold=True)
    for i, rr in enumerate(rows, start=2):
        for j, v in enumerate(rr, start=1):
            ws.cell(i, j).value = v
    ws.freeze_panes = "A2"
    for j in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(j)].width = 16
    ws.column_dimensions["L"].width = 22
    return ws

def main(xlsx_path: str, out_path: str | None = None, sheet_name: str | None = None,
         progress_every: int = 200000):
    wb = openpyxl.load_workbook(xlsx_path)
    cfg = ensure_config_sheet(wb)
    C, step, P, sat, g, target_routes, target_speeds = read_config(cfg)

    ws = find_data_sheet(wb, preferred_name=sheet_name)
    col = header_map(ws)

    candidates = []
    for route, speed, r in iter_rows(ws, col["路線番号"], col["系統速度"]):
        if route is None:
            continue
        if target_routes is not None and int(route) not in target_routes:
            continue
        if target_speeds is not None and float(speed) not in target_speeds:
            continue
        candidates.append((int(route), float(speed), r))

    total_cases = len(candidates)
    print(f"CONFIG loaded: C={C}, step={step}, P={P}, sat={sat}, g={g}", flush=True)
    print(f"Total cases to run: {total_cases}", flush=True)

    summary_rows = []

    for idx, (route, speed, r) in enumerate(candidates, start=1):
        print(f"\nCase {idx}/{total_cases}: route={route} speed={speed:g}", flush=True)
        tau01, tau12, tau23 = get_tau_from_row(ws, r, col["系統速度"], col["link(0,1)"], col["link(1,2)"], col["link(2,3)"])
        top10 = topk_offsets(tau01, tau12, tau23, C, step, P, sat, g, k=10, progress_every=progress_every)

        write_best_to_main(ws, r, col, top10[0])
        detail = make_detail_sheet(wb, route, speed, top10, C, step, P, sat, g)

        for rank, row in enumerate(top10, start=1):
            dtotal, x1, x2, x3, d0, d1, d2, d3 = row
            summary_rows.append([route, speed, rank, dtotal, x1, x2, x3, d0, d1, d2, d3, detail])

        print(f"done: route={route} speed={speed:g} best_d={top10[0][0]:.6g}", flush=True)

    make_summary_sheet(wb, summary_rows)

    if out_path is None:
        p = Path(xlsx_path)
        out_path = str(p.with_name(p.stem + "_out.xlsx"))

    wb.save(out_path)
    print(f"\nSaved: {out_path}", flush=True)

if __name__ == "__main__":
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="input xlsx (e.g. experiment.xlsx)")
    ap.add_argument("--out", default=None, help="output xlsx (default: input_stem_out.xlsx)")
    ap.add_argument("--sheet", default=None, help="data sheet name (optional). If omitted, auto-detect.")
    ap.add_argument("--progress-every", type=int, default=200000,
                    help="print inner-loop progress every N combos (default: 200000). Set 0 to disable.")
    args = ap.parse_args()
    main(args.xlsx, args.out, args.sheet, progress_every=args.progress_every)
